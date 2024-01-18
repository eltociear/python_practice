import openpyxl as op

# 既存のExcelファイルを開く
filename = ''
wb = op.load_workbook(filename, data_only=True)
ws = wb['結果']

# シート内のデータを取得
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6):
    # 行のデータを格納するリスト
    row_data = []
    for cell in row:
        # 合計列のセルが数式の場合、計算結果を取得する
        if cell.column == 6 and cell.data_type == 'f':
            row_data.append(cell.value)
        else:
            row_data.append(cell.value)
    print(tuple(row_data))
